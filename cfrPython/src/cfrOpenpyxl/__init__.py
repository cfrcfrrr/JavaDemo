#-*-coding:gbk-*-    #���ó�gbk�Ϳ����������д����
#openpyxl��������Ϻ�����д�ĺܶస���Ѿ���Ч�ˣ����ƿ���±Ƚ�Ƶ��������
import openpyxl, time, datetime
from openpyxl.utils.cell import get_column_letter
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from openpyxl.comments import Comment
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, NamedStyle
from openpyxl.styles import colors
from copy import copy
from openpyxl.cell import Cell
from openpyxl.styles.fills import FILL_SOLID
from openpyxl.styles.fills import Fill
everyBomColNum = 6

# class PdmBom():
#     #��ôָ������ı������ͣ�
#     def __init__(self, bomStr, cnDesStr, usDesStr, numInt, locStr, typeStr):
#         self.bom = bomStr
#         self.cnDes = cnDesStr
#         self.usDes = usDesStr
#         self.num = numInt
#         self.loc = locStr
#         self.type = typeStr
#         self.sonObj = []
#         self.rowNum = 2
#         self.colNum = 1
#     
#     def depth(self):
#         #�����Ǵ��˱��룬���ǲ�֪����ôͨ�����뷴�ҵ����󣬲�֪����û������java����Ļ���
#         self.addExcel()
#         for sonIndex in range(len(self.sonObj)):
#             if sonIndex == 0:
#                 self.sonObj[sonIndex].rowNum = self.rowNum
#             else:
#                 global deptestRowNum
#                 self.sonObj[sonIndex].rowNum = deptestRowNum + 1
#             self.sonObj[sonIndex].colNum = self.colNum + everyBomColNum
#             # ����Ӧ�ô��������ôͨ�������bomֵ���ҵ����� 
#             self.sonObj[sonIndex].depth()
#     
#     def addExcel(self):
#         global fontObj
#         activeSheet.cell(row=self.rowNum, column=self.colNum).value = self.bom
#         activeSheet.cell(row=self.rowNum, column=self.colNum+1).value = self.cnDes
#         activeSheet.cell(row=self.rowNum, column=self.colNum+2).value = self.usDes
#         activeSheet.cell(row=self.rowNum, column=self.colNum+3).value = self.num
#         activeSheet.cell(row=self.rowNum, column=self.colNum+4).value = self.loc
#         activeSheet.cell(row=self.rowNum, column=self.colNum+5).value = self.type
#         global deptestRowNum
#         if self.rowNum > deptestRowNum:
#             deptestRowNum = self.rowNum
#         if activeSheet.cell(row=1, column=self.colNum).value == None:
#             activeSheet.cell(row=1, column=self.colNum).value = "����"
#             activeSheet.cell(row=1, column=self.colNum+1).value = "��������"
#             activeSheet.cell(row=1, column=self.colNum+2).value = "Ӣ������"
#             activeSheet.cell(row=1, column=self.colNum+3).value = "����"
#             activeSheet.cell(row=1, column=self.colNum+4).value = "λ��"
#             activeSheet.cell(row=1, column=self.colNum+5).value = "����"

if __name__ == '__main__':
    #ģ������
#     topFatherBomObj = PdmBom('02301246', '3U ����', '3U Whole Machine', 1, '', 'PH')
#     bom02351379 = PdmBom('02351379', '��Դ', 'Power Supply', 2, 'PSU0, PSU1', 'AI')
#     bom03021234 = PdmBom('03021234', '��Դ��ģ��', 'Power Supply Son', 1, '', 'AI')
#     bom03057769 = PdmBom('03057769', '3U ���ذ�', '3U Board', 2, '', 'AI')
#     bom03037653 = PdmBom('03037653', '3U ���ذ�2', '3U Board2', 1, '', 'PH')
#     bom02301UWW = PdmBom('02301UWW', '�ڴ�', 'memory', 16, '', 'WS')
#     bom03021235 = PdmBom('03021235', '��Դ��ģ��', 'Power Supply Son', 1, '', 'PH')
#     bom03024BGY = PdmBom('03024BGY', '3U ���ذ�3', '3U Board3', 1, '', 'AI')
#     bom02301BUW = PdmBom('02301BUW', 'ϵͳ��', 'System Disk', 2, 'FANH-J11', 'AI')
#     topFatherBomObj.sonObj.append(bom02351379)
#     topFatherBomObj.sonObj.append(bom03057769)
#     bom02351379.sonObj.append(bom03021234)
#     bom03057769.sonObj.append(bom03037653)
#     bom03057769.sonObj.append(bom02301UWW)
#     bom02301UWW.sonObj.append(bom03021235)
#     bom03037653.sonObj.append(bom03024BGY)
#     bom03037653.sonObj.append(bom02301BUW)

#     global deptestRowNum
#     deptestRowNum = 0
    
    #������WorkBook����
    wb = openpyxl.Workbook()
    #��ȡ��ǰ���
    ws = wb.get_active_sheet()
    #�޸ı���
    ws.title = 'test1'
    
    #�����±�
    test2ws = wb.create_sheet(title="test2")
    test3ws = wb.create_sheet("test3")
    
    #���ر�
    loadWorkBook = load_workbook(filename = 'tmp.xlsx')
    loadWorkSheet = loadWorkBook['Sheet1']
    
    #�����Զ�ʶ���������ͣ�Ĭ��ΪTrue
    wb.guess_types = True
    ws['A1'] = '3.14%'
    print(ws['A1'].value) #0.031400000000000004
    print(ws['A1'].number_format) #0%
    wb.guess_types = False
    ws['A2'] = '3.14%'
    print(ws['A2'].value) #3.14%
    print(ws['A2'].number_format) #General
    
    #ʹ�ù�ʽ
    ws['A3'] = "=SUM(1, 1)"
    print(ws['A3'].value) #������ӡ������ֵ���ǹ�ʽ��=SUM(1, 1)���� ʵ���ϱ������ʾ����Ϊ2
    
    #�ϲ���ȡ���ϲ���Ԫ��
    ws['B1'] = 'bbb111'
    ws['B2'] = 'bbb222'
    ws['C1'] = 'ccc111'
    ws['C2'] = 'ccc222'
    ws.merge_cells('B1:C2')
    print(ws['B1'].value) #bbb111
    print(ws['B2'].value) #None
    print(ws['C1'].value) #None
    print(ws['C2'].value) #None
    ws.unmerge_cells('B1:C2')
    print(ws['B1'].value) #bbb111
    print(ws['B2'].value) #None
    print(ws['C1'].value) #None
    print(ws['C2'].value) #None
    
    ws['B1'] = 'bbb111'
    ws['B2'] = 'bbb222'
    ws['C1'] = 'ccc111'
    ws['C2'] = 'ccc222'
    ws.merge_cells(start_row=1, start_column=2, end_row=2, end_column = 3)
    ws.unmerge_cells(start_row=1, start_column=2, end_row=2, end_column = 3)
    
    #����ͼƬ
#     tmpImg = Image('tmpImg.png')
#     ws.add_image(tmpImg, 'D1')

    #�۵��У��Ҷ���֪��excel�����⹦��
#     ws.column_dimensions.group('A', 'C', hidden=True)
#     ws.column_dimensions.group('A', 'C', hidden=False)������չ������
    
    #���ע�ͣ����Ǳ������û����������֪��ʲô���
#     commentObj = ws['B1'].comment
#     commentObj = Comment('This is the comment text', 'Comment Author')
#     print(commentObj.text) #This is the comment text
#     print(commentObj.author) #Comment Author
    #���ַ�ʽ����
    commentObj2 = Comment('Text3', 'Author2')
    ws['B1'].comment = commentObj2
    
    #��ʽ
    #Ĭ��ֵ
    #����
#   font = Font(name='Calibri', #����
#                  size=11,    #�ֺ�
#                  bold=False,    #�Ӵ�
#                  italic=False,    #��б
#                  vertAlign=None,    #����Ĵ�ֱ���뷽ʽ��������Ϊ'baseline', 'superscript', 'subscript'���ֱ�Ϊ���С����ϺͿ���
#                  underline='none',    #�»���
#                  strike=False,    #ɾ����
#                  color='FF000000')    #������ɫ
    #���䣿
#  fill = PatternFill(fill_type=None,
#                  start_color='FFFFFFFF',
#                  end_color='FF000000')
    #�߿�
#  border = Border(left=Side(border_style=None,
#                            color='FF000000'),
#                  right=Side(border_style=None,
#                             color='FF000000'),
#                  top=Side(border_style=None,
#                           color='FF000000'),
#                  bottom=Side(border_style=None,
#                              color='FF000000'),
#                  diagonal=Side(border_style=None,
#                                color='FF000000'),
#                  diagonal_direction=0,
#                  outline=Side(border_style=None,
#                               color='FF000000'),
#                  vertical=Side(border_style=None,
#                                color='FF000000'),
#                  horizontal=Side(border_style=None,
#                                 color='FF000000')
#                 )
    #����
#  alignment=Alignment(horizontal='general',    #ˮƽ���룬'general', 'right', 'justify', 'centerContinuous', 'distributed', 'fill', 'left', 'center'
#                      vertical='bottom',  #��ֱ����, 'top', 'bottom', 'center', 'distributed', 'justify'
#                      text_rotation=0,    #��ת
#                      wrap_text=False,    #�Զ�����
#                      shrink_to_fit=False,    #��С������ƥ�䵥Ԫ���С
#                      indent=0
#                      readingOrder #���ַ���
#                        
#                         )    #��������
#  number_format = 'General'    #���õ�Ԫ��ķ��࣬@-�ı�
#  protection = Protection(locked=True,
#                          hidden=False)

    #Font����������ʽ��color-��ɫ��italic-��б��name-���塢size-�ֺš�bold-�Ӵ֡�underline-�»���
    ftObj = Font(color=colors.RED, italic=True, name='Arial', size=14, bold=True, underline="single", strike=True)
    ws['D1'] = 'Hello World!'
    ws['D1'].font = ftObj
    #��ֱ������
    ws['D2'] = 'DDD222'
    #��һ��������ɫ�ķ�ʽ
    ws['D2'].font = Font(color="FFBB00")
    
    #Font�����ܹ�����
    ws['D3'] = 'ddd333'
    ftObj2 = copy(ftObj)
    ftObj2.name = 'Tahoma'
    ws['D3'].font = ftObj2
    
    #��ʽ�ܹ����ø����л����У�����������ܺ����ϣ����������ݲ���Ч���Կո��ڹر��ļ������Ч������������ȥ��ֵ��ʱ���ֱ��ʹ��Ĭ�ϵĸ�ʽ
#     colObj = ws.column_dimensions['D'] # �����������и�������0.0
#     colObj.font = Font(size=30)
#     rowObj = ws.row_dimensions[1]
#     rowObj.font = Font(underline="single")
#     #�ر��ļ�
#     saveFilename = time.strftime('%Y-%m-%d_%H-%M-%S', time.localtime(time.time())) + '.xlsx'
#     wb.save(saveFilename)
#     wb = load_workbook(filename = saveFilename)
#     ws = wb.get_active_sheet()
#     ws['D4'] = 'ddd444'
    
    #��֪���⼸�������õģ������������иߺ��п�������û����Ч
#     ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
#     ws.page_setup.paperSize = ws.PAPERSIZE_TABLOID
    ws.page_setup.fitToHeight = 30
    ws.page_setup.fitToWidth = 30

    #������ʽ
    #����������ʽ
    highlightNS = NamedStyle(name="highlight")
    highlightNS.fone = Font(bold=True, size=20)
    bd = Side(style='thick', color='000000')
    highlightNS.border = Border(left=bd, top=bd, right=bd, bottom=bd)
    #���������������ʽ
    wb.add_named_style(highlightNS)
    #��Ԫ��ͨ����ʽ����������ʽ
    ws['E1'].style = highlightNS
    ws['E1'] = 'EEE111'
    #��Ԫ��ͨ����ʽ����������ʽ
    ws['E2'].style = 'highlight'
    ws['E2'] = 'EEE222'
    ws['E3'] = 'EEE333'
    #����ʹ��ϵͳĬ�ϵ���ʽ��ϵͳ�Դ������£�
    # ��Normal��-���桢��Good��-�á���Bad��-���Neutral��-����
    # ��Calculation��-���㡢��Check Cell��-��鵥Ԫ�񡢡�Explanatory Text��-�������ı�����Warning Text��-�����ı�����Linked Cell��-���ӵ�Ԫ��
    # ��Output��-�������Input��-���롢��Note��-ע��
    # ��Title��-���⡢��Headline 1��-����1����Headline 2��-����2����Headline 3��-����3����Headline 4��-����4����Total��-����
    # ��Accent1��-ǿ��������ɫ1����20 % - Accent1��-20%ǿ��������ɫ1����40 % - Accent1��-40%ǿ��������ɫ1����60 % - Accent1��-60%ǿ��������ɫ1
    # ��Accent2��-ǿ��������ɫ2����20 % - Accent2��-20%ǿ��������ɫ2����40 % - Accent2��-40%ǿ��������ɫ2����60 % - Accent2��-60%ǿ��������ɫ2
    # ��Accent3��-ǿ��������ɫ3����20 % - Accent3��-20%ǿ��������ɫ3����40 % - Accent3��-40%ǿ��������ɫ3����60 % - Accent3��-60%ǿ��������ɫ3
    # ��Accent4��-ǿ��������ɫ4����20 % - Accent4��-20%ǿ��������ɫ4����40 % - Accent4��-40%ǿ��������ɫ4����60 % - Accent4��-60%ǿ��������ɫ4
    # ��Accent5��-ǿ��������ɫ5����20 % - Accent5��-20%ǿ��������ɫ5����40 % - Accent5��-40%ǿ��������ɫ5����60 % - Accent5��-60%ǿ��������ɫ5
    # ��Accent6��-ǿ��������ɫ6����20 % - Accent6��-20%ǿ��������ɫ6����40 % - Accent6��-40%ǿ��������ɫ6����60 % - Accent6��-60%ǿ��������ɫ6
    #��Percent��-�ٷֱȡ���Comma��-���ҡ���Comma [0]��-����[0]����Currency��-ǧλ�ָ��Currency [0]��-ǧλ�ָ�[0]
    # ��Hyperlink������Followed Hyperlink������Pandas��
    
    #�����п�
    ws['F1'] = 'Life is short, I use python.'
    ws.column_dimensions['F'].width = 60
    
    ws['G1'] = '01234567'
    ws['G1'].number_format = '@'
    print(ws['G1'].number_format)
    #�Ӷ��㿪ʼ������ȱ����ڵ���
#     topFatherBomObj.depth()
    
    #��������
#     ws.freeze_panes = 'A2'
#     #���õ�Ԫ����ʽ
#     for rowI in range(1, ws.max_row + 1):
#         for colJ in range(1, ws.max_column + 1):
#             #horizontal����������롢���С��Ҷ��룻vertical�������϶��롢���С��¶��룻wrap_text�������Ƿ��Զ�����
#             ws.cell(row=rowI, column=colJ).alignment = alignment.Alignment(horizontal='left', vertical='center', wrap_text=True)

    #����������
    # Available properties for worksheets
    # ��enableFormatConditionsCalculation��
    # ��filterMode��
    # ��published��
    # ��syncHorizontal��
    # ��syncRef��
    # ��syncVertical��
    # ��transitionEvaluation��
    # ��transitionEntry��
    # ��tabColor��
    # 
    # 
    # Available fields for page setup properties
    # 
    # ��autoPageBreaks�� ��fitToPage��
    # 
    # 
    # Available fields for outlines
    # ��applyStyles��
    # ��summaryBelow��
    # ��summaryRight��
    # ��showOutlineSymbols��
    ws['A1'] = '����'
    print(ws['E5'].col_idx)
    print(ws['A1'].encoding)

    #���浽�ļ���
#     filename = topFatherBomObj.bom + time.strftime('_%Y-%m-%d_%H-%M-%S', time.localtime(time.time())) + '.xlsx'
    filename = time.strftime('%Y-%m-%d_%H-%M-%S', time.localtime(time.time())) + '.xlsx'
    wb.save(filename)
    print("Pass")